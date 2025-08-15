# import openpyxl
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import Font, Alignment, Border, Side, numbers as fmt
# from io import BytesIO
# from datetime import datetime
# from django.http import HttpResponse


# class OpenpxlHelper:
#     def __init__(
#         self,
#         sheet_name: str,
#         data: list[list],
#         filename="report",
#         setting: list[str] = None,
#         title: str = None,
#         options: list[str] = None,
#         auto_width=True,
#     ):
#         self.sheet_name = sheet_name
#         self.data = data
#         self.setting = setting
#         self.auto_width = auto_width
#         self.title = title
#         self.options = options or []
#         self.filename = f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
#         self.wb = openpyxl.Workbook()
#         self.ws = self.wb.active
#         self.ws.title = self.sheet_name

#         self._generate()

#     def _generate(self):
#         current_row = 1
#         total_columns = len(self.data[0]) if self.data else 1

#         if self.title:
#             self.ws.merge_cells(
#                 start_row=current_row,
#                 start_column=1,
#                 end_row=current_row,
#                 end_column=total_columns,
#             )
#             cell = self.ws.cell(row=current_row, column=1, value=self.title)
#             cell.font = Font(size=14, bold=True)
#             cell.alignment = Alignment(horizontal="center")
#             current_row += 1

#         self.ws.row_dimensions[current_row].height = 8
#         current_row += 1

#         for option in self.options:
#             self.ws.merge_cells(
#                 start_row=current_row,
#                 start_column=1,
#                 end_row=current_row,
#                 end_column=total_columns,
#             )
#             cell = self.ws.cell(row=current_row, column=1, value=option)
#             cell.font = Font(italic=True)
#             cell.alignment = Alignment(horizontal="left")
#             current_row += 1

#         if self.options:
#             self.ws.row_dimensions[current_row].height = 8
#             current_row += 1

#         for row_idx, row in enumerate(self.data, start=current_row):
#             for col_idx, value in enumerate(row, start=1):
#                 cell = self.ws.cell(row=row_idx, column=col_idx, value=value)

#                 thin_border = Border(
#                     left=Side(style="thin"),
#                     right=Side(style="thin"),
#                     top=Side(style="thin"),
#                     bottom=Side(style="thin"),
#                 )
#                 cell.border = thin_border

#                 if row_idx == current_row:
#                     cell.font = Font(bold=True)
#                     cell.alignment = Alignment(horizontal="center")
#                 else:
#                     if self.setting and col_idx <= len(self.setting):
#                         parts = self.setting[col_idx - 1].split()
#                         for part in parts:
#                             if part.startswith("w-"):
#                                 try:
#                                     width = int(part.replace("w-", ""))
#                                     col_letter = get_column_letter(col_idx)
#                                     self.ws.column_dimensions[col_letter].width = width
#                                 except ValueError:
#                                     pass
#                             elif part == "text-right":
#                                 cell.alignment = Alignment(horizontal="right")
#                             elif part == "text-left":
#                                 cell.alignment = Alignment(horizontal="left")
#                             elif part == "text-center":
#                                 cell.alignment = Alignment(horizontal="center")
#                             elif part == "date":
#                                 cell.number_format = "DD-MM-YYYY"
#                                 if isinstance(value, str):
#                                     try:
#                                         parsed_date = datetime.strptime(
#                                             value, "%Y-%m-%d"
#                                         ).date()
#                                         cell.value = parsed_date
#                                     except ValueError:
#                                         pass
#                             elif part == "number":
#                                 cell.number_format = "#,##0"

#         if self.auto_width and not self.setting:
#             for col in self.ws.columns:
#                 max_length = max(
#                     (len(str(cell.value)) if cell.value else 0) for cell in col
#                 )
#                 col_letter = get_column_letter(col[0].column)
#                 self.ws.column_dimensions[col_letter].width = max_length + 2

#     @property
#     def excel_file(self) -> BytesIO:
#         output = BytesIO()
#         self.wb.save(output)
#         output.seek(0)
#         return output

#     @property
#     def response(self) -> HttpResponse:
#         response = HttpResponse(
#             self.excel_file.getvalue(),
#             content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         )
#         response["Content-Disposition"] = f'attachment; filename="{self.filename}"'
#         return response


import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, numbers as fmt
from io import BytesIO
from datetime import datetime
from django.http import HttpResponse


class OpenpxlHelper:
    def __init__(
        self,
        data: list[list],
        filename="report",
        setting: list[str] = None,
        title: str = None,
        options: list[str] = None,
        auto_width=True,
    ):
        # self.sheet_name = sheet_name
        self.data = data
        self.setting = setting or []
        self.auto_width = auto_width
        self.title = title
        self.options = options or []
        self.filename = f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        # self.ws.title = self.sheet_name

        self._generate()

    def _generate(self):
        current_row = 1
        total_columns = len(self.data[0]) if self.data else 1

        if self.title:
            self.ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=total_columns,
            )
            cell = self.ws.cell(row=current_row, column=1, value=self.title)
            cell.font = Font(size=14, bold=True)
            cell.alignment = Alignment(horizontal="center")
            current_row += 1

        self.ws.row_dimensions[current_row].height = 8
        current_row += 1

        for option in self.options:
            self.ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=total_columns,
            )
            cell = self.ws.cell(row=current_row, column=1, value=option)
            cell.font = Font(italic=True)
            cell.alignment = Alignment(horizontal="left")
            current_row += 1

        if self.options:
            self.ws.row_dimensions[current_row].height = 8
            current_row += 1

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Data for aggregation
        column_values = {i: [] for i in range(len(self.data[0]))}

        # Write data rows (including header)
        for row_idx, row in enumerate(self.data, start=current_row):
            for col_idx, value in enumerate(row, start=1):
                cell = self.ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                if row_idx == current_row:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                else:
                    if self.setting and col_idx <= len(self.setting):
                        parts = self.setting[col_idx - 1].split()

                        for part in parts:
                            if part.startswith("w-"):
                                try:
                                    width = int(part.replace("w-", ""))
                                    col_letter = get_column_letter(col_idx)
                                    self.ws.column_dimensions[col_letter].width = width
                                except ValueError:
                                    pass
                            elif part == "text-right":
                                cell.alignment = Alignment(horizontal="right")
                            elif part == "text-left":
                                cell.alignment = Alignment(horizontal="left")
                            elif part == "text-center":
                                cell.alignment = Alignment(horizontal="center")
                            elif part == "date":
                                cell.number_format = "DD-MM-YYYY"
                                if isinstance(value, str):
                                    try:
                                        parsed_date = datetime.strptime(
                                            value, "%Y-%m-%d"
                                        ).date()
                                        cell.value = parsed_date
                                    except ValueError:
                                        pass
                            elif part == "number":
                                cell.number_format = "#,##0"

                        # Simpan nilai numerik untuk agregasi
                        if any(k in parts for k in ("sum", "avg", "count")):
                            try:
                                column_values[col_idx - 1].append(float(value))
                            except:
                                pass

        # Tambahkan baris agregasi jika ada setting "sum", "avg", "count"
        if any(any(k in s for k in ("sum", "avg", "count")) for s in self.setting):
            agg_row = current_row + len(self.data)
            for col_idx in range(1, total_columns + 1):
                parts = (
                    self.setting[col_idx - 1].split()
                    if col_idx - 1 < len(self.setting)
                    else []
                )
                value = ""
                if "sum" in parts:
                    value = sum(column_values[col_idx - 1])
                elif "avg" in parts:
                    values = column_values[col_idx - 1]
                    value = sum(values) / len(values) if values else 0
                elif "count" in parts:
                    value = len(column_values[col_idx - 1])

                if value != "":
                    cell = self.ws.cell(row=agg_row, column=col_idx, value=value)
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0.00" if "avg" in parts else "#,##0"

    @property
    def excel_file(self) -> BytesIO:
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output

    @property
    def response(self) -> HttpResponse:
        response = HttpResponse(
            self.excel_file.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{self.filename}"'
        return response


"""
### Example Data
data = [
    ["No", "Nama", "Tanggal Lahir", "Gaji"],
    ["1", "Joe", "1988-04-24", 250000],
    ["2", "Daisy", "1984-04-24", 150000],
]

setting = [
    "w-5 text-center",
    "w-40",
    "w-25 date",
    "w-20 number sum"
]

### Example Usage
response_type = query_params.get("response_type")
if response_type in ["xlsx", "pdf"]:
    export_data = [["No", "Nama"]]
    export_setting = ["w-5 text-center", "w-40"]
    for idx, item in enumerate(result["data"], start=1):
        export_data.append([idx, item.get("name", "")])

    export_kwargs = {
        "data": export_data,
        "setting": export_setting,
        "filename": "author_grade",
        "title": "Data Author Grade",
    }

    if response_type == "pdf":
        export_kwargs["options"] = ["Cluster: Sinom", "Blok: B1"]

    helper_class = {
        "xlsx": OpenpxlHelper,
        "pdf": PdfTableHelper,
    }.get(response_type)

    return helper_class(**export_kwargs).response
"""
