from datetime import datetime
from openpyxl import load_workbook


class ExcelUtils:
    @staticmethod
    def read(file, schema=None, exclude=[0]):
        wb = load_workbook(filename=file.file, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None

        headers = [str(h).strip() for i, h in enumerate(rows[0]) if i not in exclude]

        data = []
        for row in rows[1:]:
            row = [
                str(v).strip() if v is not None else ""
                for i, v in enumerate(row)
                if i not in exclude
            ]
            item = dict(zip(headers, row))
            if schema:
                valid, casted = ExcelUtils.validate_and_cast_row(item, schema)
                if not valid:
                    return None
                data.append(casted)
            else:
                data.append(item)

        return data

    @staticmethod
    def validate_and_cast_row(row, schema):
        casted = {}

        date_formats = [
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%Y/%m/%d",
            "%m/%d/%Y",
            "%m-%d-%Y",
            "%d.%m.%Y",
            "%Y.%m.%d",
            "%d/%m/%y",
            "%d-%m-%y",
            "%m/%d/%y",
            "%m-%d-%y",
        ]

        for field, rule in schema.items():
            raw_value = row.get(field)
            if raw_value is None or raw_value.strip() == "":
                if rule.get("required"):
                    return False, None
                casted[field] = None
                continue

            try:
                if rule["type"] == "int":
                    casted[field] = int(raw_value)
                elif rule["type"] == "float":
                    casted[field] = float(raw_value)
                elif rule["type"] == "date":
                    if isinstance(raw_value, datetime):
                        casted[field] = raw_value.date()
                    else:
                        parsed = None
                        for fmt in date_formats:
                            try:
                                parsed = datetime.strptime(raw_value, fmt).date()
                                break
                            except ValueError:
                                continue
                        casted[field] = parsed
                elif rule["type"] == "enum":
                    if raw_value not in rule.get("choices", []):
                        return False, None
                    casted[field] = raw_value
                elif rule["type"] == "phone":
                    digits = "".join(filter(str.isdigit, raw_value))
                    if not digits.startswith(("08", "628")):
                        return False, None
                    if len(digits) < 9 or len(digits) > 13:
                        return False, None
                    if digits.startswith("08"):
                        digits = "62" + digits[1:]
                    casted[field] = digits
                else:
                    casted[field] = raw_value
            except Exception:
                return False, None

        return True, casted


"""
Usage
schema = {
    "LEVEL": {"type": "int", "required": True},
    "NAMA": {"type": "string", "required": True},
    "TGL_LAHIR": {"type": "date", "required": False},
    "NAMA": {"type": "enum", "choices": ["VIII A", "VIII B"], "required": True},
    "NO. ORANG TUA": {"type": "phone", "required": True}
}

excel_data = ExcelUtils.read(validated_data.get("file"), schema=schema)
"""
